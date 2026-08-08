import io
import math

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .. import models
from ..deps import get_db, get_current_user

router = APIRouter(prefix="/forms", tags=["export"])

# ============================================================
# STYLING HELPERS
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="2563EB")
BAND_FILL = PatternFill("solid", fgColor="F1F5F9")
TITLE_FILL = PatternFill("solid", fgColor="DBEAFE")

WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)

THIN_SIDE = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = WHITE_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def _style_body(ws, first_row, last_row, ncols):
    for r in range(first_row, last_row + 1):
        band = (r - first_row) % 2 == 1
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if band and cell.fill.patternType is None:
                cell.fill = BAND_FILL


def _style_title_row(ws, row, ncols):
    cell = ws.cell(row=row, column=1)
    cell.font = BOLD
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER
    for c in range(2, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = TITLE_FILL
        cell.border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


def _auto_width(ws, max_width=60, min_width=8):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            for line in str(cell.value).split("\n"):
                longest = max(longest, len(line))
        ws.column_dimensions[letter].width = max(min_width, min(longest + 2, max_width))


def _apply_row_heights(ws, first_row, last_row, min_height=18):
    for r in range(first_row, last_row + 1):
        max_lines = 1
        for cell in ws[r]:
            if cell.value is None:
                continue
            width = ws.column_dimensions[cell.column_letter].width or 8.43
            chars_per_line = max(8, int(width) - 2)
            lines = 0
            for part in str(cell.value).split("\n"):
                lines += max(1, math.ceil(len(part) / chars_per_line))
            max_lines = max(max_lines, lines)
        ws.row_dimensions[r].height = max(min_height, max_lines * 15 + 4)


# ============================================================
# DATA HELPERS
# ============================================================
def _fmt_datetime(dt):
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _duration_str(sub):
    if not (sub.started_at and sub.submitted_at):
        return "-"
    total_secs = max(0, int((sub.submitted_at - sub.started_at).total_seconds()))
    mins, secs = divmod(total_secs, 60)
    return f"{mins}m {secs}s" if mins > 0 else f"{secs}s"


def _correct_keys(question):
    return {o.label for o in question.options if o.is_correct}


def _is_graded(question):
    return len(_correct_keys(question)) > 0


def _user_selected(ans):
    if not ans:
        return []
    if ans.answer_options:
        return list(ans.answer_options)
    if ans.answer_text:
        return [ans.answer_text]
    return []


def _is_correct(question, ans):
    keys = _correct_keys(question)
    return len(keys) > 0 and keys == set(_user_selected(ans))


def _answer_value(ans):
    if not ans:
        return ""
    if ans.answer_text:
        return ans.answer_text
    if ans.answer_options:
        return ", ".join(ans.answer_options)
    if ans.file_url:
        return ans.file_url
    return ""


def _score_percent(question_map, sub):
    total_gradable = sum(1 for q in question_map.values() if _is_graded(q))
    if total_gradable == 0:
        return None
    answers = {a.question_id: a for a in sub.answers}
    correct = sum(
        1
        for q in question_map.values()
        if _is_graded(q) and _is_correct(q, answers.get(q.id))
    )
    return round((correct / total_gradable) * 100)


# ============================================================
# EXPORT ENDPOINT
# ============================================================
@router.get("/{form_id}/export")
def export_submissions_to_excel(
    form_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    form = db.query(models.Form).filter(models.Form.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form tidak ditemukan")
    if form.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Bukan form milikmu")

    questions = (
        db.query(models.Question)
        .filter(models.Question.form_id == form_id)
        .order_by(models.Question.order_index)
        .all()
    )
    submissions = db.query(models.Submission).filter(models.Submission.form_id == form_id).all()
    submissions.sort(key=lambda s: (s.submitted_at is None, s.submitted_at or s.started_at))

    question_map = {q.id: q for q in questions}
    has_graded = any(_is_graded(q) for q in questions)

    wb = Workbook()

    # --------------------------------------------------------
    # SHEET 1: REKAP RESPONDEN
    # --------------------------------------------------------
    ws = wb.active
    ws.title = "Rekap Responden"

    rekap_headers = ["No", "Nama Pengisi", "Email", "Waktu Mulai", "Waktu Submit", "Durasi"]
    if has_graded:
        rekap_headers.append("Skor /100")
    rekap_headers.append("Status")

    ws.append(rekap_headers)

    for idx, sub in enumerate(submissions, start=1):
        name = sub.user.full_name if sub.user else "Responden (User)"
        email = sub.user.email if sub.user else ""
        is_completed = bool(sub.submitted_at)

        row = [
            idx,
            name,
            email,
            _fmt_datetime(sub.started_at),
            _fmt_datetime(sub.submitted_at),
            _duration_str(sub),
        ]
        if has_graded:
            score = _score_percent(question_map, sub)
            row.append(f"{score}/100" if score is not None else "-")
        row.append("Selesai" if is_completed else "Proses")
        ws.append(row)

    _style_header_row(ws, 1, len(rekap_headers))
    _style_body(ws, 2, ws.max_row, len(rekap_headers))
    _auto_width(ws)
    _apply_row_heights(ws, 2, ws.max_row)
    ws.freeze_panes = "A2"

    # --------------------------------------------------------
    # SHEET 2: DETAIL JAWABAN
    # --------------------------------------------------------
    ws = wb.create_sheet("Detail Jawaban")

    detail_headers = ["No", "Nama Pengisi", "Email", "Waktu Submit"] + [
        f"{i + 1}. {q.label}" for i, q in enumerate(questions)
    ]
    ws.append(detail_headers)

    for idx, sub in enumerate(submissions, start=1):
        answers = {a.question_id: a for a in sub.answers}
        name = sub.user.full_name if sub.user else "Responden (User)"
        email = sub.user.email if sub.user else ""

        row = [idx, name, email, _fmt_datetime(sub.submitted_at)]
        for q in questions:
            ans = answers.get(q.id)
            value = _answer_value(ans)
            if value and _is_graded(q):
                value = ("✓ " if _is_correct(q, ans) else "✗ ") + value
            row.append(value)
        ws.append(row)

    _style_header_row(ws, 1, len(detail_headers))
    _style_body(ws, 2, ws.max_row, len(detail_headers))
    _auto_width(ws)
    _apply_row_heights(ws, 2, ws.max_row)
    ws.freeze_panes = "D2"

    # --------------------------------------------------------
    # SHEET 3: ANALISIS JAWABAN
    # --------------------------------------------------------
    ws = wb.create_sheet("Analisis Jawaban")
    total_respondents = len(submissions)
    option_types = {"single_choice", "checkbox", "dropdown"}
    ncols = 4
    row = 1

    for idx, q in enumerate(questions, start=1):
        _style_title_row(ws, row, ncols)
        ws.cell(row=row, column=1).value = f"{idx}. {q.label} ({q.type})"
        row += 1

        answered = 0
        for sub in submissions:
            ans = {a.question_id: a for a in sub.answers}.get(q.id)
            if _answer_value(ans):
                answered += 1

        if q.type in option_types and q.options:
            headers = ["Opsi", "Jumlah Responden", "Persentase", "Keterangan"]
            for c, h in enumerate(headers, start=1):
                ws.cell(row=row, column=c).value = h
            _style_header_row(ws, row, ncols)
            row += 1

            for opt in q.options:
                count = 0
                for sub in submissions:
                    ans = {a.question_id: a for a in sub.answers}.get(q.id)
                    if ans and opt.label in _user_selected(ans):
                        count += 1
                percent = round((count / answered) * 100) if answered else 0
                keterangan = "✓ Kunci" if opt.is_correct else ""
                ws.append([opt.label, count, f"{percent}%", keterangan])
                row += 1

            ws.append(["Total responden menjawab", answered, "", ""])
            ws.append(["Total responden tidak menjawab", total_respondents - answered, "", ""])
            row += 2
        else:
            headers = ["Jumlah Menjawab", "Total Responden", "Persentase", "Contoh Jawaban"]
            for c, h in enumerate(headers, start=1):
                ws.cell(row=row, column=c).value = h
            _style_header_row(ws, row, ncols)
            row += 1

            percent = round((answered / total_respondents) * 100) if total_respondents else 0
            samples = []
            for sub in submissions:
                ans = {a.question_id: a for a in sub.answers}.get(q.id)
                value = _answer_value(ans)
                if value and len(samples) < 3:
                    samples.append(value)
            ws.append([answered, total_respondents, f"{percent}%", "\n".join(samples)])
            row += 2

    _style_body(ws, 2, ws.max_row, ncols)
    _auto_width(ws)
    _apply_row_heights(ws, 2, ws.max_row)
    ws.freeze_panes = "A2"

    # --------------------------------------------------------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{form.slug}-hasil.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
