import asyncio
import unittest
from io import BytesIO
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import models
from app.database import Base
from app.main import app
from app.routers.export import export_submissions_to_excel


class RouteRegistrationTests(unittest.TestCase):
    def test_template_and_question_routes_are_registered(self):
        paths = {route.path for route in app.routes if hasattr(route, "path")}

        self.assertIn("/templates/{template_id}", paths)
        self.assertIn("/forms/{form_id}/questions", paths)
        self.assertIn("/templates/{template_id}/questions", paths)
        self.assertIn("/questions/{question_id}", paths)
        self.assertIn("/questions/{question_id}/options", paths)
        self.assertIn("/options/{option_id}", paths)

    def test_auth_me_supports_put_for_profile_update(self):
        put_routes = [
            route for route in app.routes
            if getattr(route, "path", None) == "/auth/me" and "PUT" in getattr(route, "methods", set())
        ]

        self.assertTrue(put_routes, "PUT /auth/me should be registered")

    def test_submission_result_and_flag_cheated_routes_are_registered(self):
        paths = {route.path for route in app.routes if hasattr(route, "path")}

        self.assertIn("/submissions/{submission_id}/result", paths)
        self.assertIn("/submissions/{submission_id}/flag-cheated", paths)

    def test_regenerate_join_token_route_is_registered(self):
        paths = {route.path for route in app.routes if hasattr(route, "path")}

        self.assertIn("/forms/{form_id}/regenerate-join-token", paths)

    def test_export_includes_only_submitted_answers_and_score(self):
        engine = create_engine("sqlite://")
        SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
        Base.metadata.create_all(bind=engine)

        with SessionLocal() as db:
            owner = models.User(
                full_name="Owner",
                email="owner@example.com",
                password_hash="hash",
            )
            db.add(owner)
            db.commit()
            db.refresh(owner)

            form = models.Form(
                owner_id=owner.id,
                title="Quiz",
                slug="quiz-export",
                status=models.FormStatus.published,
                accept_responses=True,
            )
            db.add(form)
            db.commit()
            db.refresh(form)

            question = models.Question(
                form_id=form.id,
                type=models.QuestionType.single_choice,
                label="Capital of Indonesia",
                is_required=True,
                order_index=1,
            )
            db.add(question)
            db.commit()
            db.refresh(question)

            option = models.QuestionOption(
                question_id=question.id,
                label="Jakarta",
                value="Jakarta",
                is_correct=True,
                order_index=1,
            )
            db.add(option)
            db.commit()

            submitted = models.Submission(
                form_id=form.id,
                user_id=owner.id,
                started_at=datetime.utcnow(),
                submitted_at=datetime.utcnow(),
            )
            db.add(submitted)
            db.commit()
            db.refresh(submitted)

            submitted_answer = models.Answer(
                submission_id=submitted.id,
                question_id=question.id,
                answer_text="Jakarta",
            )
            db.add(submitted_answer)
            db.commit()

            draft = models.Submission(
                form_id=form.id,
                user_id=owner.id,
                started_at=datetime.utcnow(),
                submitted_at=None,
            )
            db.add(draft)
            db.commit()

            response = export_submissions_to_excel(form.id, db=db, current_user=owner)

            async def _read_stream(iterable):
                chunks = []
                async for chunk in iterable:
                    chunks.append(chunk)
                return b"".join(chunks)

            workbook_bytes = asyncio.run(_read_stream(response.body_iterator))
            workbook = load_workbook(filename=BytesIO(workbook_bytes))

            recap_rows = list(workbook["Rekap Responden"].iter_rows(values_only=True))
            self.assertIn("Skor /100", recap_rows[0])
            self.assertEqual(len(recap_rows) - 1, 1, "Draft submissions should not be exported")

            detail_rows = list(workbook["Detail Jawaban"].iter_rows(values_only=True))
            self.assertIn("Skor /100", detail_rows[0])
            self.assertTrue(any("Jakarta" in str(cell) for row in detail_rows for cell in row))


if __name__ == "__main__":
    unittest.main()
